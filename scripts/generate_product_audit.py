"""
Generate product audit Excel file.
Sheet1 — products with no image
Sheet2 — products with no description
"""

import urllib.request
import urllib.parse
import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TOKEN = os.environ.get("LIGHTSPEED_API_TOKEN", "")
BASE_URL = os.environ.get("LIGHTSPEED_BASE_URL", "https://justthetipcigars.retail.lightspeed.app/api/2.0")

SWAG_TYPES = {
    'Short Sleeve Tee', 'Hats', 'Beanie', 'Hoodie', 'Trucker',
    'Long Sleeve Tee', 'Shirts', 'Flex Fit', 'Polo',
}

CIGAR_SIZES = [
    'robusto', 'toro', 'gordo', 'corona', 'lancero', 'churchill', 'belicoso',
    'panatela', 'lonsdale', 'perfecto', 'presidente', 'torpedo', 'figurado',
    'petit', 'gran', 'double', 'magnum', 'salomon', 'croqueta', 'puritos',
    'short', 'gigante', 'hermoso', 'julieta', 'dalia',
]

def is_cigar_size(value):
    v = value.lower()
    return any(s in v for s in CIGAR_SIZES) or bool(re.search(r'\d+\s*[xX×]\s*\d+', v))

def fetch_json(url):
    req = urllib.request.Request(url, headers={
        'Authorization': f'Bearer {TOKEN}',
        'Content-Type': 'application/json',
    })
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode())

def fetch_all_products():
    products = []
    after = None
    while True:
        url = f"{BASE_URL}/products?page_size=250"
        if after:
            url += f"&after={after}"
        data = fetch_json(url)
        page = data.get('data', [])
        for p in page:
            if p.get('deleted_at') or p.get('is_active') is not True or p.get('price_including_tax') is None:
                continue
            if p.get('name') == 'Discount':
                continue
            price = p.get('price_including_tax', 0)
            if not price or price <= 0:
                continue

            variant_options = p.get('variant_options', []) or []
            size = next((v['value'] for v in variant_options if v.get('name') == 'Size'), '')

            type_name = (p.get('type') or {}).get('name', '')
            if SWAG_TYPES and type_name in SWAG_TYPES:
                category = 'JTTC Swag'
            else:
                cat = p.get('product_category') or {}
                category = cat.get('name', '') or type_name

            brand_obj = p.get('brand') or {}
            brand = brand_obj.get('name', '')

            # Image detection
            image_url = p.get('image_url', '')
            images = p.get('images', []) or []
            has_image = bool(
                (image_url and 'placeholder' not in image_url) or
                any(i.get('url', '') and 'placeholder' not in i.get('url', '') for i in images)
            )

            description = (p.get('description') or '').strip()

            products.append({
                'name': p.get('name', ''),
                'brand': brand,
                'category': category,
                'size': size,
                'sku': p.get('sku', ''),
                'price': price,
                'has_image': has_image,
                'has_description': bool(description),
            })

        if len(page) == 250:
            version = data.get('version') or {}
            after = version.get('max')
            if not after:
                break
        else:
            break
    return products

def group_by_name(products):
    groups = {}
    for p in products:
        name = p['name']
        if name not in groups:
            groups[name] = {
                'name': name,
                'brand': p['brand'],
                'category': p['category'],
                'skus': set(),
                'has_image': False,
                'has_description': False,
            }
        g = groups[name]
        if p['sku']:
            g['skus'].add(p['sku'])
        if p['has_image']:
            g['has_image'] = True
        if p['has_description']:
            g['has_description'] = True
    return sorted(groups.values(), key=lambda g: g['name'])

HEADER_FILL = PatternFill(fill_type='solid', fgColor='2D2D2D')
HEADER_FONT = Font(bold=True, color='F5F0E8', size=10)
ALT_FILL = PatternFill(fill_type='solid', fgColor='1A1A1A')

def style_sheet(ws, headers, rows):
    ws.freeze_panes = 'A2'

    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.row_dimensions[1].height = 18

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if fill:
                cell.fill = fill

    # Auto-width
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value or ''
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

def main():
    print("Fetching products from Lightspeed...")
    products = fetch_all_products()
    print(f"  {len(products)} active variants fetched")

    groups = group_by_name(products)
    print(f"  {len(groups)} unique products grouped")

    no_image = [g for g in groups if not g['has_image']]
    no_desc  = [g for g in groups if not g['has_description']]
    print(f"  {len(no_image)} products missing images")
    print(f"  {len(no_desc)} products missing descriptions")

    wb = Workbook()

    # ── Sheet 1: No Image ──
    ws1 = wb.active
    ws1.title = 'Missing Images'
    ws1.sheet_properties.tabColor = 'C0392B'

    headers1 = ['Product Name', 'Brand', 'Category', 'SKUs']
    rows1 = [
        [g['name'], g['brand'], g['category'], ', '.join(sorted(g['skus']))]
        for g in no_image
    ]
    style_sheet(ws1, headers1, rows1)

    # ── Sheet 2: No Description ──
    ws2 = wb.create_sheet('Missing Descriptions')
    ws2.sheet_properties.tabColor = 'E67E22'

    headers2 = ['Product Name', 'Brand', 'Category', 'SKUs']
    rows2 = [
        [g['name'], g['brand'], g['category'], ', '.join(sorted(g['skus']))]
        for g in no_desc
    ]
    style_sheet(ws2, headers2, rows2)

    # ── Sheet 3: Missing Both ──
    no_image_names = {g['name'] for g in no_image}
    no_desc_names  = {g['name'] for g in no_desc}
    both = [g for g in groups if g['name'] in no_image_names and g['name'] in no_desc_names]
    print(f"  {len(both)} products missing both image and description")

    ws3 = wb.create_sheet('Missing Both')
    ws3.sheet_properties.tabColor = '8E44AD'

    headers3 = ['Product Name', 'Brand', 'Category', 'SKUs']
    rows3 = [
        [g['name'], g['brand'], g['category'], ', '.join(sorted(g['skus']))]
        for g in both
    ]
    style_sheet(ws3, headers3, rows3)

    out_path = os.path.join(os.path.dirname(__file__), '..', 'product_audit.xlsx')
    out_path = os.path.normpath(out_path)
    wb.save(out_path)
    print(f"\nSaved: {out_path}")

if __name__ == '__main__':
    main()
